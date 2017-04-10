from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from pathlib import Path
def getexcel(path):
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    row = ws.max_row
    col = min(7, ws.max_column)
    if row < 32:
        print(path)
    works = dict()
    for i in range(2, row+1):
        workname = ws.cell(row=i, column=1).value
        works[workname] = list()
        for j in range(2, col):
            value = ws.cell(row=i, column=j).value
            if value is not None:
                works[workname].append(value)
            else:
                works[workname].append(0)
        works[workname].append(sum(works[workname]))
    return works

def maxminofdict(md):
    mdlist = list()
    for k,v in md.items():
        mdlist.append((v,k))
    mdlist.sort()
    return (mdlist[-1][1], mdlist[0][1])

path = Path('/media/nouzan/DATA/Onedrive/文档/中山大学/网宣部/评分表/')
markfiles = list(path.glob('*.xlsx'))
marks = dict()
for file in markfiles:
    if '评分表' in file.stem:
        marks[file.stem.replace('评分表-', '')] = getexcel(str(file))
workmarks = dict()
for pw in marks:
    for workname in marks[pw]:
        if workname not in workmarks:
            workmarks[workname] = dict()
        workmarks[workname][pw] = marks[pw][workname][-1]
total = dict()
total_raw = dict()
for workname in workmarks:
    total[workname] = 0
    for pw in workmarks[workname]:
        total[workname] = total[workname] + float(workmarks[workname][pw])
    total_raw[workname] = total[workname]
    for pw in maxminofdict(workmarks[workname]):
        total[workname] = total[workname] - float(workmarks[workname][pw])
    total[workname] = total[workname]/float(len(workmarks[workname])-2)
    total_raw[workname] = total_raw[workname]/float(len(workmarks[workname]))
totalist = list()
for k, v in total.items():
    totalist.append((v, k))
totalist.sort(reverse = True)
print("\n去掉最高分和最低分:")
count = 0
for v, k in totalist:
    count = count + 1
    print(str(count)+'.', k, round(v*100)/100)

totalist = list()
for k, v in total_raw.items():
    totalist.append((v, k))
totalist.sort(reverse = True)
print("\n原始分数：")
count = 0
for v, k in totalist:
    count = count + 1
    print(str(count)+'.', k, round(v*100)/100)
wb = Workbook()
wb_filename = 'result.xlsx'
ws = wb.worksheets[0]
ws.title = '评分总览'
rowx = 1
for name in total:
    rowx = rowx + 1
    ws.cell(row=rowx, column=1).value = name
    ws.cell(row=rowx, column=2).value = total[name]
    ws.cell(row=rowx, column=3).value = total_raw[name]
ws.cell(row=1, column=1).value = '作品'
ws.cell(row=1, column=2).value = '去掉最高最低分的平均分'
ws.cell(row=1, column=3).value = '平均分'
wb.save(str(path) +'/' +wb_filename)
